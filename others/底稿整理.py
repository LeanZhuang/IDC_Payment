from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import os
import glob
import pandas as pd


# 设置下载文件夹路径和目标文件夹路径
download_folder = "/Users/zhuangyuhao/Downloads"
desktop_folder = "/Users/zhuangyuhao/Desktop"

# 获取下载文件夹内最新的xls文件
latest_file = max(glob.glob(os.path.join(download_folder, "*.xls")), key=os.path.getctime)

# 读取xls文件
df = pd.read_excel(latest_file)

# 构建目标文件路径，将文件后缀名改为xlsx
target_file = os.path.join(desktop_folder, os.path.basename(latest_file).replace(".xls", ".xlsx"))

# 将数据保存为xlsx文件
df.to_excel(target_file, index=False)

# 清理内存以释放资源 (可选)
del df


# 获取桌面上最新的 xlsx 文件路径
desktop_path = os.path.expanduser("~/Desktop")
files = [f for f in os.listdir(desktop_path) if f.endswith(".xlsx")]
latest_file = max(files, key=lambda x: os.path.getctime(os.path.join(desktop_path, x)))
file_path = os.path.join(desktop_path, latest_file)

# 加载工作簿
workbook = load_workbook(file_path)
sheet = workbook.active

# 定义值和公式所在的区域范围
last_row = sheet.max_row

start_row = last_row + 1
end_row = last_row + 4
start_col = 'E'
end_col = 'G'

word_list = ['非带宽小计', '带宽小计', '合计', '付款申请与对账单核对']

formula_list1 = [f'=SUMIF(M1:M{last_row},"机架",F1:F{last_row})+SUMIF(M1:M{last_row},"IP",F1:F{last_row})',
                 f'=SUMIF(M1:M{last_row},"端口组",F1:F{last_row})',
                 f'=F{last_row + 1}+F{last_row + 2}']

formula_list2 = [f'=F{last_row + 1}-VLOOKUP("合计",Q:R,2,FALSE)',
                 f'=F{last_row + 2}-VLOOKUP("合计",H:I,2,FALSE)']



# 循环遍历要输入值的列表
for index, value in enumerate(word_list):
    row = last_row + index + 1  # 计算要输入行的行号
    cell = 'E{}'.format(row)  # 构建单元格地址
    sheet[cell] = value  # 输入值到单元格

# 循环遍历要输入公式的范围每一行
for index, value in enumerate(formula_list1):
    row = last_row + index + 1  # 计算要输入行的行号
    cell = 'F{}'.format(row)  # 构建单元格地址
    sheet[cell] = value  # 输入值到单元格

# 循环遍历要输入公式的范围每一行
for index, value in enumerate(formula_list2):
    row = last_row + index + 1  # 计算要输入行的行号
    cell = 'G{}'.format(row)  # 构建单元格地址
    sheet[cell] = value  # 输入值到单元格


fill = PatternFill(fill_type='solid', fgColor='ACD6FF')


for row in range(last_row + 1, last_row + 5):
    for col in ['E', 'F', 'G']:
        cell = '{}{}'.format(col, row)
        # sheet[cell].font = font
        sheet[cell].fill = fill


# 保存工作簿
workbook.save('./middle_data/中间底稿.xlsx')
