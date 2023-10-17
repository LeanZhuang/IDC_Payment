from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
import model.prepare_accured_data as prepare_accured_data

# 加载底稿
accured_table = prepare_accured_data.prepare_accured_data()
accured_table = accured_table.active

# 加载匹配到的预提
wb1 = load_workbook('middle_data/非带宽.xlsx')
ws1 = wb1.active
ws1.insert_rows(0)
ws1.insert_rows(0)

wb2 = load_workbook('middle_data/带宽.xlsx')
ws2 = wb2.active
ws2.insert_rows(0)
ws2.insert_rows(0)

# 合并三个Excel文件的数据
merged_data =  list(accured_table.values) + list(ws1.values) + list(ws2.values)

# 创建空工作簿
merged_wb = load_workbook('others/empty_file.xlsx')

# 设置工作表为合并后数据的工作表
merged_ws = merged_wb.active
merged_ws.title = "Merged Data"


# 添加合并后的数据
for row in merged_data:
    merged_ws.append(row)


last_row = merged_ws.max_row
row = last_row + 1
cell = f'E{row}'
merged_ws[cell] = f'=(D{row}-C{row})/D{row}'

cell = f'F{row}'
merged_ws[cell] = f'=AVERAGE(C{row},D{row})'

cell = f'I{row}'
merged_ws[cell] = f'=G{row}*H{row}'


row = accured_table.max_row + ws1.max_row
cell = f'R{row}'
merged_ws[cell] = f'=SUM(R{accured_table.max_row + 4}:R{row - 1})'


# 配置颜色
blue = PatternFill(fill_type='solid', fgColor='ACD6FF')
gray = PatternFill(fill_type='solid', fgColor='DFDFDF')


for row in range(1, 2):
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K','L','M','N','O','P','Q','R','S','T','U','V','W']:
        cell = '{}{}'.format(col, row)
        merged_ws[cell].fill = blue

for row in range(accured_table.max_row - 3, accured_table.max_row + 1):
    for col in ['E', 'F', 'G']:
        cell = '{}{}'.format(col, row)
        merged_ws[cell].fill = gray

for row in range(accured_table.max_row + 3, accured_table.max_row + 4):
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y']:
        cell = '{}{}'.format(col, row)
        merged_ws[cell].fill = blue

for row in range(accured_table.max_row + 3 + ws1.max_row, accured_table.max_row + 4 + ws1.max_row):
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R','S','T','U','V','W','X','Y','Z','AA','AB','AC']:
        cell = '{}{}'.format(col, row)
        merged_ws[cell].fill = blue

for row in range(accured_table.max_row + ws1.max_row + ws2.max_row, accured_table.max_row + ws1.max_row + ws2.max_row + 1):
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        cell = '{}{}'.format(col, row)
        merged_ws[cell].fill = gray


# 保存合并后的Excel文件
merged_wb.save('/Users/zhuangyuhao/Desktop/【底稿】.xlsx')
