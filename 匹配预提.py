import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import accrued_expense_data

# 加载预提表
no_bandwidth_list, bandwidth_list = accrued_expense_data.prepare_accured()
# try:
#     expense_2305_bandwidth = pd.read_pickle('2305 带宽.pkl')
#     expense_2305_no_bandwidth = pd.read_pickle('2305 非带宽.pkl')

#     expense_2306_bandwidth = pd.read_pickle('2306 带宽.pkl')
#     expense_2306_no_bandwidth = pd.read_pickle('2306 非带宽.pkl')

#     expense_2307_bandwidth = pd.read_pickle('2307 带宽.pkl')
#     expense_2307_no_bandwidth = pd.read_pickle('2307 非带宽.pkl')

#     expense_2308_bandwidth = pd.read_pickle('2308 带宽.pkl')
#     expense_2308_no_bandwidth = pd.read_pickle('2308 非带宽.pkl')

#     expense_2309_bandwidth = pd.read_pickle('2309 带宽.pkl')
#     expense_2309_no_bandwidth = pd.read_pickle('2309 非带宽.pkl')

# except FileNotFoundError:
#     # 2306
#     expense_2305_bandwidth = pd.read_excel('/Users/zhuangyuhao/Documents/Fileport/预算表/2305 带宽.xlsx')
#     expense_2305_no_bandwidth = pd.read_excel('/Users/zhuangyuhao/Documents/Fileport/预算表/2305 非带宽.xlsx')
#     expense_2305_bandwidth.to_pickle('2305 带宽.pkl')
#     expense_2305_no_bandwidth.to_pickle('2305 非带宽.pkl')

#     # 2306
#     expense_2306_bandwidth = pd.read_excel('/Users/zhuangyuhao/Documents/Fileport/预算表/2306 带宽.xlsx')
#     expense_2306_no_bandwidth = pd.read_excel('/Users/zhuangyuhao/Documents/Fileport/预算表/2306 非带宽.xlsx')
#     expense_2306_bandwidth.to_pickle('2306 带宽.pkl')
#     expense_2306_no_bandwidth.to_pickle('2306 非带宽.pkl')

#     # 2307
#     expense_2307_bandwidth = pd.read_excel('/Users/zhuangyuhao/Documents/Fileport/预算表/2307 带宽.xlsx', sheet_name='202307带宽')
#     expense_2307_no_bandwidth = pd.read_excel('/Users/zhuangyuhao/Documents/Fileport/预算表/2307 非带宽.xlsx')
#     expense_2307_bandwidth.to_pickle('2307 带宽.pkl')
#     expense_2307_no_bandwidth.to_pickle('2307 非带宽.pkl')

#     # 2308
#     expense_2308_bandwidth = pd.read_excel('/Users/zhuangyuhao/Documents/Fileport/预算表/2308 带宽.xlsx')
#     expense_2308_no_bandwidth = pd.read_excel('/Users/zhuangyuhao/Documents/Fileport/预算表/2308 非带宽.xlsx')
#     expense_2308_bandwidth.to_pickle('2308 带宽.pkl')
#     expense_2308_no_bandwidth.to_pickle('2308 非带宽.pkl')

#     # 2309
#     expense_2309_bandwidth = pd.read_excel('/Users/zhuangyuhao/Documents/Fileport/预算表/2309 带宽.xlsx')
#     expense_2309_no_bandwidth = pd.read_excel('/Users/zhuangyuhao/Documents/Fileport/预算表/2309 非带宽.xlsx')
#     expense_2309_bandwidth.to_pickle('2309 带宽.pkl')
#     expense_2309_no_bandwidth.to_pickle('2309 非带宽.pkl')


# no_bandwidth_list = [expense_2305_no_bandwidth, expense_2306_no_bandwidth, expense_2308_no_bandwidth, expense_2307_no_bandwidth, expense_2309_no_bandwidth]
# bandwidth_list = [expense_2305_bandwidth, expense_2306_bandwidth, expense_2308_bandwidth, expense_2307_bandwidth, expense_2309_bandwidth]


desktop_path = os.path.expanduser("~/Desktop")
files = [f for f in os.listdir(desktop_path) if f.endswith(".xlsx")]
latest_file = max(files, key=lambda x: os.path.getctime(os.path.join(desktop_path, x)))
file_path = os.path.join(desktop_path, latest_file)

check = pd.read_excel(file_path)


contract_id = check['合同编号'].to_list()
contract_id = list(set(contract_id))

no_bandwidth_need = pd.DataFrame([])
bandwidth_need = pd.DataFrame([])


for id in contract_id:
    for database in no_bandwidth_list:
        database = database[database['当前计提合同'] == id]
        no_bandwidth_need = pd.concat([no_bandwidth_need, database])

for id in contract_id:
    for database in bandwidth_list:
        database = database[database['当前计提合同'] == id]
        bandwidth_need = pd.concat([bandwidth_need, database])


period_list = check['费用表月份'].to_list()

period_list = list(map(str, period_list))

period_list = [period.replace('-', '') for period in period_list]
period_list = list(set(period_list))

bandwidth_need['费用期间'] = bandwidth_need['费用期间'].astype(str)
no_bandwidth_need['费用期间'] = no_bandwidth_need['费用期间'].astype(str)


bandwidth_need = bandwidth_need[bandwidth_need['费用期间'].isin(period_list)]
no_bandwidth_need = no_bandwidth_need[no_bandwidth_need['费用期间'].isin(period_list)]


# 保存中间表
bandwidth_need.to_excel('/Users/zhuangyuhao/Downloads/带宽.xlsx', index=False)
no_bandwidth_need.to_excel('/Users/zhuangyuhao/Downloads/非带宽.xlsx', index=False)


# 格式化带宽预提

# 加载工作簿
workbook = load_workbook('/Users/zhuangyuhao/Downloads/带宽.xlsx')
sheet = workbook.active

# 定义值和公式所在的区域范围
last_row = sheet.max_row

col_list = {'A':'地点', 'B':'地点', 'C':'SYS统计',
            'D':'运营商统计', 'E':'差异率', 'F':'中值',
            'G':'结算流量', 'H':'计费单位', 'I':'结算'}

fill = PatternFill(fill_type='solid', fgColor='ACD6FF')

# 循环遍历要输入值的列表
for key, value in col_list.items():
    row = last_row + 2  # 计算要输入行的行号
    cell = f'{key}{row}'  # 构建单元格地址
    sheet[cell] = value  # 输入值到单元格
    sheet[cell].fill = fill


row = last_row + 3
cell = f'E{row}'
sheet[cell] = f'=(D{row}-C{row})/D{row}'

cell = f'F{row}'
sheet[cell] = f'=AVERAGE(C{row},D{row})'

cell = f'I{row}'
sheet[cell] = f'=G{row}*H{row}'


workbook.save('/Users/zhuangyuhao/Downloads/带宽.xlsx')


# 格式化非带宽预提

# 加载工作簿
workbook = load_workbook('/Users/zhuangyuhao/Downloads/非带宽.xlsx')
sheet = workbook.active

# 定义值和公式所在的区域范围
last_row = sheet.max_row


fill = PatternFill(fill_type='solid', fgColor='FFFF6F')

row = last_row + 1
cell = f'R{row}'
sheet[cell] = f'=SUM(R2:R{row-1})'
sheet[cell].fill = fill


cell = f'Q{row}'
sheet[cell] = f'合计'
sheet[cell].fill = fill


workbook.save('/Users/zhuangyuhao/Downloads/非带宽.xlsx')
